"""
python src/live_capture_cicflow_xgb_excel.py `
  --interface 5 `
  --dumpcap-exe "C:\Program Files\Wireshark\dumpcap.exe" `
  --cicflowmeter-exe "C:\Users\abdul\miniconda3\envs\cicflow\Scripts\cicflowmeter.exe" `
  --window-seconds 30 `
  --attack-label "Kali Live Test" `
  --excel outputs\live_attack_results.xlsx `
  --min-attack-flows 1 `
  --max-windows 1
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import time
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook


# =========================================================
# Paths
# =========================================================
PROJECT_ROOT = Path(".")
LIVE_DIR = PROJECT_ROOT / "live_capture"
PCAP_DIR = LIVE_DIR / "pcaps"
FLOW_RAW_DIR = LIVE_DIR / "flows_raw"
FLOW_READY_DIR = LIVE_DIR / "flows_model_ready"
PRED_DIR = LIVE_DIR / "predictions"
SUMMARY_DIR = LIVE_DIR / "summaries"

PROCESSED_DIR = PROJECT_ROOT / "dataset" / "processed"
MODEL_DIR = PROJECT_ROOT / "models" / "baseline_xgb"

FEATURES_JSON = PROCESSED_DIR / "feature_columns_binary.json"
MEDIANS_JSON = PROCESSED_DIR / "median_imputer_binary.json"
LABEL_MAP_JSON = PROCESSED_DIR / "label_mapping_binary.json"
THRESHOLD_JSON = MODEL_DIR / "selected_threshold.json"
MODEL_PATH = MODEL_DIR / "xgb_baseline_model.joblib"

DEFAULT_EXCEL = PROJECT_ROOT / "outputs" / "live_attack_results.xlsx"


# =========================================================
# Column mapping from CICFlowMeter CSV -> training schema
# =========================================================
COLUMN_MAP = {
    "dst_port": "destination_port",
    "tot_fwd_pkts": "total_fwd_packets",
    "tot_bwd_pkts": "total_backward_packets",
    "totlen_fwd_pkts": "total_length_of_fwd_packets",
    "totlen_bwd_pkts": "total_length_of_bwd_packets",
    "fwd_pkt_len_max": "fwd_packet_length_max",
    "fwd_pkt_len_min": "fwd_packet_length_min",
    "fwd_pkt_len_mean": "fwd_packet_length_mean",
    "fwd_pkt_len_std": "fwd_packet_length_std",
    "bwd_pkt_len_max": "bwd_packet_length_max",
    "bwd_pkt_len_min": "bwd_packet_length_min",
    "bwd_pkt_len_mean": "bwd_packet_length_mean",
    "bwd_pkt_len_std": "bwd_packet_length_std",
    "flow_byts_s": "flow_bytes_per_s",
    "flow_pkts_s": "flow_packets_per_s",
    "fwd_iat_tot": "fwd_iat_total",
    "bwd_iat_tot": "bwd_iat_total",
    "fwd_header_len": "fwd_header_length",
    "bwd_header_len": "bwd_header_length",
    "fwd_pkts_s": "fwd_packets_per_s",
    "bwd_pkts_s": "bwd_packets_per_s",
    "pkt_len_min": "min_packet_length",
    "pkt_len_max": "max_packet_length",
    "pkt_len_mean": "packet_length_mean",
    "pkt_len_std": "packet_length_std",
    "pkt_len_var": "packet_length_variance",
    "fin_flag_cnt": "fin_flag_count",
    "syn_flag_cnt": "syn_flag_count",
    "rst_flag_cnt": "rst_flag_count",
    "psh_flag_cnt": "psh_flag_count",
    "ack_flag_cnt": "ack_flag_count",
    "urg_flag_cnt": "urg_flag_count",
    "ece_flag_cnt": "ece_flag_count",
    "cwr_flag_count": "cwe_flag_count",
    "pkt_size_avg": "average_packet_size",
    "fwd_seg_size_avg": "avg_fwd_segment_size",
    "bwd_seg_size_avg": "avg_bwd_segment_size",
    "subflow_fwd_pkts": "subflow_fwd_packets",
    "subflow_fwd_byts": "subflow_fwd_bytes",
    "subflow_bwd_pkts": "subflow_bwd_packets",
    "subflow_bwd_byts": "subflow_bwd_bytes",
    "init_fwd_win_byts": "init_win_bytes_forward",
    "init_bwd_win_byts": "init_win_bytes_backward",
    "fwd_act_data_pkts": "act_data_pkt_fwd",
    "fwd_seg_size_min": "min_seg_size_forward",
}


# =========================================================
# Helpers
# =========================================================
def load_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def make_json_safe(obj):
    if isinstance(obj, dict):
        return {str(k): make_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]
    if isinstance(obj, tuple):
        return [make_json_safe(v) for v in obj]
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, np.floating):
        return float(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    return obj


def save_json(path: Path, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(make_json_safe(data), f, indent=2)


def now_text() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def ensure_dirs() -> None:
    for d in [PCAP_DIR, FLOW_RAW_DIR, FLOW_READY_DIR, PRED_DIR, SUMMARY_DIR]:
        d.mkdir(parents=True, exist_ok=True)


def clear_old_capture_contents() -> None:
    LIVE_DIR.mkdir(parents=True, exist_ok=True)
    for d in [PCAP_DIR, FLOW_RAW_DIR, FLOW_READY_DIR, PRED_DIR, SUMMARY_DIR]:
        if d.exists():
            shutil.rmtree(d)
    ensure_dirs()


def ensure_excel(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append([
        "Attack itself",
        "Classification",
        "Window file",
        "Flows total",
        "Flows predicted ATTACK",
        "Attack rate",
        "Max attack probability",
        "Mean attack probability",
        "Threshold used",
        "Processed at",
    ])
    wb.save(path)


def append_excel_row(path: Path, row: list[object]) -> None:
    ensure_excel(path)
    wb = load_workbook(path)
    ws = wb["Results"]
    ws.append(row)
    wb.save(path)


def run_command(cmd: list[str], description: str) -> None:
    print(f"[RUN] {description}")
    print(" ".join(f'"{c}"' if " " in str(c) else str(c) for c in cmd))
    completed = subprocess.run(cmd)
    if completed.returncode != 0:
        raise RuntimeError(f"{description} failed with exit code {completed.returncode}")


def capture_one_window(
    dumpcap_exe: str,
    interface: str,
    capture_filter: str,
    window_seconds: int,
    pcap_path: Path,
) -> None:
    cmd = [
        dumpcap_exe,
        "-i", interface,
        "-f", capture_filter,
        "-F", "pcap",
        "-a", f"duration:{window_seconds}",
        "-w", str(pcap_path),
    ]
    run_command(cmd, f"capture window -> {pcap_path.name}")


def run_cicflowmeter(
    cicflowmeter_exe: str,
    pcap_path: Path,
    csv_path: Path,
) -> None:
    cmd = [
        cicflowmeter_exe,
        "-f", str(pcap_path),
        "-c", str(csv_path),
    ]
    run_command(cmd, f"cicflowmeter -> {csv_path.name}")

    if not csv_path.exists() or csv_path.stat().st_size == 0:
        raise RuntimeError(f"CICFlowMeter did not create a usable CSV: {csv_path}")

def coalesce_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    if not df.columns.duplicated().any():
        return df

    merged = {}
    ordered_cols = []

    for col in pd.unique(df.columns):
        col_data = df.loc[:, df.columns == col]

        if isinstance(col_data, pd.Series):
            merged[col] = col_data
        else:
            merged[col] = col_data.bfill(axis=1).iloc[:, 0]

        ordered_cols.append(col)

    return pd.DataFrame(merged, columns=ordered_cols)

def adapt_cicflowmeter_csv(
    input_csv: Path,
    output_csv: Path,
    feature_cols: list[str],
) -> tuple[pd.DataFrame, dict]:
    df = pd.read_csv(input_csv, low_memory=False)

    df = df.rename(columns=COLUMN_MAP)
    df = coalesce_duplicate_columns(df)
    
    keep_meta = []
    for col in ["src_ip", "dst_ip", "src_port", "destination_port", "protocol", "timestamp"]:
        if col in df.columns:
            keep_meta.append(col)

    originally_present = set(df.columns)

    for col in feature_cols:
        if col not in df.columns:
            df[col] = pd.NA
###
    out_cols = []
    seen = set()

    for col in keep_meta + feature_cols:
        if col not in seen:
            out_cols.append(col)
            seen.add(col)

    df = df[out_cols]
###

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_csv, index=False)

    info = {
        "input_columns_count": len(originally_present),
        "mapped_or_existing_feature_count": sum(1 for c in feature_cols if c in originally_present or c in COLUMN_MAP.values()),
        "missing_features_added": [c for c in feature_cols if c not in originally_present],
    }
    return df, info


def run_model_on_ready_df(
    df_ready: pd.DataFrame,
    feature_cols: list[str],
    medians: dict[str, float],
    model,
    attack_threshold: float,
    attack_class_id: int,
    benign_class_id: int,
) -> tuple[pd.DataFrame, dict]:
    X = df_ready[feature_cols].copy()

    bad_tokens = {
        "Infinity": np.nan,
        "INFINITY": np.nan,
        "inf": np.nan,
        "Inf": np.nan,
        "-Infinity": np.nan,
        "-inf": np.nan,
        "-Inf": np.nan,
        "NaN": np.nan,
        "nan": np.nan,
        "": np.nan,
        " ": np.nan,
    }
#
    for col in feature_cols:
        col_data = X[col]

        if isinstance(col_data, pd.DataFrame):
            col_data = col_data.bfill(axis=1).iloc[:, 0]

        col_data = col_data.replace(bad_tokens)
        col_data = pd.to_numeric(col_data, errors="coerce")
        X[col] = col_data
    #
    X = X.replace([np.inf, -np.inf], np.nan)
    X = X.fillna(medians)
    X = X.fillna(0)

    proba = model.predict_proba(X)
    classes = [int(x) for x in model.classes_]
    attack_idx = classes.index(attack_class_id)

    attack_probs = proba[:, attack_idx]
    pred_ids = np.where(attack_probs >= attack_threshold, attack_class_id, benign_class_id)
    pred_names = np.where(pred_ids == attack_class_id, "ATTACK", "BENIGN")

    out_df = df_ready.copy()
    out_df["attack_probability"] = attack_probs
    out_df["predicted_target_id"] = pred_ids
    out_df["predicted_target_name"] = pred_names
    out_df["used_attack_threshold"] = attack_threshold

    total_flows = int(len(out_df))
    attack_flows = int((pred_ids == attack_class_id).sum())
    benign_flows = int((pred_ids == benign_class_id).sum())
    attack_rate = float(attack_flows / total_flows) if total_flows else 0.0
    max_attack_prob = float(np.max(attack_probs)) if total_flows else 0.0
    mean_attack_prob = float(np.mean(attack_probs)) if total_flows else 0.0

    summary = {
        "flows_total": total_flows,
        "flows_predicted_attack": attack_flows,
        "flows_predicted_benign": benign_flows,
        "attack_rate": attack_rate,
        "max_attack_probability": max_attack_prob,
        "mean_attack_probability": mean_attack_prob,
        "threshold_used": attack_threshold,
    }
    return out_df, summary


def window_classification(summary: dict, min_attack_flows: int) -> str:
    return "ATTACK" if int(summary["flows_predicted_attack"]) >= min_attack_flows else "BENIGN"


# =========================================================
# Main
# =========================================================
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Capture live PCAP windows, convert with CICFlowMeter, adapt to training schema, run XGBoost, append Excel results."
    )
    parser.add_argument("--interface", required=True, help="Capture interface number or name for dumpcap")
    parser.add_argument("--dumpcap-exe", default="dumpcap", help="Path to dumpcap.exe or just 'dumpcap' if on PATH")
    parser.add_argument(
        "--cicflowmeter-exe",
        default="cicflowmeter",
        help="Path to cicflowmeter.exe. Example: C:\\Users\\abdul\\miniconda3\\envs\\cicflow\\Scripts\\cicflowmeter.exe",
    )
    parser.add_argument("--window-seconds", type=int, default=30, help="Capture duration per window")
    parser.add_argument("--capture-filter", default="tcp or udp or icmp", help="BPF capture filter")
    parser.add_argument("--attack-label", default="Live Attack Window", help="Text to write in the 'Attack itself' column")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help="Excel file path")
    parser.add_argument("--min-attack-flows", type=int, default=1, help="Window is ATTACK if at least this many flows are predicted ATTACK")
    parser.add_argument("--max-windows", type=int, default=0, help="0 = run until Ctrl+C, otherwise stop after this many windows")
    args = parser.parse_args()

    clear_old_capture_contents()
    excel_path = Path(args.excel)
    ensure_excel(excel_path)

    feature_cols = load_json(FEATURES_JSON)["feature_columns"]
    medians = load_json(MEDIANS_JSON)
    label_map = load_json(LABEL_MAP_JSON)
    attack_threshold = float(load_json(THRESHOLD_JSON)["best_threshold"])
    model = joblib.load(MODEL_PATH)

    attack_class_id = int(label_map["ATTACK"])
    benign_class_id = int(label_map["BENIGN"])

    print("[INFO] Old live_capture contents deleted.")
    print(f"[INFO] Window size: {args.window_seconds} seconds")
    print(f"[INFO] Excel output: {excel_path}")
    print("[INFO] Press Ctrl+C to stop.\n")

    window_num = 1

    try:
        while True:
            if args.max_windows > 0 and window_num > args.max_windows:
                break

            stamp = time.strftime("%Y%m%d_%H%M%S")
            base_name = f"window_{window_num:03d}_{stamp}"

            pcap_path = PCAP_DIR / f"{base_name}.pcap"
            flow_raw_csv = FLOW_RAW_DIR / f"{base_name}.csv"
            flow_ready_csv = FLOW_READY_DIR / f"{base_name}_model_ready.csv"
            pred_csv = PRED_DIR / f"{base_name}_predictions.csv"
            summary_json = SUMMARY_DIR / f"{base_name}_summary.json"

            capture_one_window(
                dumpcap_exe=args.dumpcap_exe,
                interface=args.interface,
                capture_filter=args.capture_filter,
                window_seconds=args.window_seconds,
                pcap_path=pcap_path,
            )

            run_cicflowmeter(
                cicflowmeter_exe=args.cicflowmeter_exe,
                pcap_path=pcap_path,
                csv_path=flow_raw_csv,
            )

            df_ready, adapt_info = adapt_cicflowmeter_csv(
                input_csv=flow_raw_csv,
                output_csv=flow_ready_csv,
                feature_cols=feature_cols,
            )

            pred_df, pred_summary = run_model_on_ready_df(
                df_ready=df_ready,
                feature_cols=feature_cols,
                medians=medians,
                model=model,
                attack_threshold=attack_threshold,
                attack_class_id=attack_class_id,
                benign_class_id=benign_class_id,
            )

            pred_df.to_csv(pred_csv, index=False)

            final_class = window_classification(pred_summary, args.min_attack_flows)

            append_excel_row(excel_path, [
                args.attack_label,
                final_class,
                pcap_path.name,
                pred_summary["flows_total"],
                pred_summary["flows_predicted_attack"],
                pred_summary["attack_rate"],
                pred_summary["max_attack_probability"],
                pred_summary["mean_attack_probability"],
                pred_summary["threshold_used"],
                now_text(),
            ])

            save_json(summary_json, {
                "attack_itself": args.attack_label,
                "classification": final_class,
                "pcap_file": str(pcap_path),
                "raw_flow_csv": str(flow_raw_csv),
                "model_ready_csv": str(flow_ready_csv),
                "prediction_csv": str(pred_csv),
                "adapt_info": adapt_info,
                "prediction_summary": pred_summary,
                "processed_at": now_text(),
            })

            print(f"[DONE] {base_name} -> {final_class}")
            print(f"       flows={pred_summary['flows_total']}, attack_flows={pred_summary['flows_predicted_attack']}, attack_rate={pred_summary['attack_rate']:.6f}\n")

            window_num += 1

    except KeyboardInterrupt:
        print("\n[INFO] Stopped by user.")

    print("[DONE] Script finished.")


if __name__ == "__main__":
    main()